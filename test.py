import tkinter as tk
from tkinter import filedialog
import openpyxl

class ExcelAdder:
    def __init__(self, master):
        self.master = master
        self.master.title("Excel Adder")
        self.master.geometry("400x400")
        self.file_path = None

        # Create buttons
        self.browse_button = tk.Button(self.master, text="Select File", command=self.select_file)
        self.browse_button.pack(pady=10)

        self.add_button = tk.Button(self.master, text="Add Data", command=self.add_data)
        self.add_button.pack(pady=10)

        # Create labels and entry fields
        self.name_label = tk.Label(self.master, text="Name:")
        self.name_label.pack()

        self.name_entry = tk.Entry(self.master)
        self.name_entry.pack()

        self.age_label = tk.Label(self.master, text="Age:")
        self.age_label.pack()

        self.age_entry = tk.Entry(self.master)
        self.age_entry.pack()

        self.hobby_label = tk.Label(self.master, text="Hobby:")
        self.hobby_label.pack()

        self.hobby_entry = tk.Entry(self.master)
        self.hobby_entry.pack()

        self.status_label = tk.Label(self.master, text="")
        self.status_label.pack()

    def select_file(self):
        # Open file dialog to select file
        self.file_path = filedialog.askopenfilename(initialdir="/", title="Select File",
                                                    filetypes=[("Excel files", "*.xlsx")])

        # Show selected file path in label
        self.status_label.config(text=f"Selected File: {self.file_path}")

    def add_data(self):
        if not self.file_path:
            self.status_label.config(text="Please select a file first")
            return

        # Open workbook and worksheet
        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active

        # Get data from entry fields
        name = self.name_entry.get()
        age = self.age_entry.get()
        hobby = self.hobby_entry.get()

        # Add data to new row
        row = [name, age, hobby]
        worksheet.append(row)

        # Save changes
        workbook.save(self.file_path)

        self.status_label.config(text="Data added successfully")

        # Show added data
        added_data_label = tk.Label(self.master, text="Added Data:")
        added_data_label.pack()

        # Get the column headings
        headings = []
        for column in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in column:
                headings.append(cell.value)

        # Get the last row (i.e. the row that was just added)
        last_row = []
        for column in worksheet.iter_cols(min_row=worksheet.max_row, max_row=worksheet.max_row):
            for cell in column:
                last_row.append(cell.value)

        # Create a label to display the added data
        added_data_str = ""
        for i in range(len(headings)):
            added_data_str += f"{headings[i]}: {last_row[i]}\n"

        added_data_value_label = tk.Label(self.master, text=added_data_str)
        added_data_value_label.pack()


if __name__ == '__main__':
    root = tk.Tk()
    app = ExcelAdder(root)
    root.mainloop()
